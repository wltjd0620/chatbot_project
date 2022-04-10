import openpyxl
import random
from openpyxl import Workbook, load_workbook
# from kobert_transformers import get_tokenizer
from kogpt2_transformers import get_kogpt2_tokenizer

#겹치는 내용들은 최초에만 주석처리를 하고 이후에는 안했습니다. 모르시면 윗부분을 봐주세요
#wellness_dialog_dataset.xlsx파일은 웰니스_대화_스크립트_데이터셋.xlsx인 걸로 추측됩니다. 코드해석에 도움이 될 겁니다

#트위터 대화 데이터(엑셀파일)의 내용을 담은 메모장파일(tweeter_dialog_data.txt)을 생성한다
def tweet_dialog_dataset():
  root_path = "../data"
  tweet_file = root_path + "/tweeter_dialog_dataset.xlsx" #데이터파일
  tweet_file_output = root_path + "/tweeter_dialog_data.txt" #옮길 메모장파일(생성할 파일)

  f = open(tweet_file_output, 'w') #tweet_file_output파일을 생성하고 쓰기모드로 설정한다 'r'면 읽기모드

  wb = load_workbook(filename=tweet_file)#tweet_file파일을 wb로 불러온다

  ws = wb[wb.sheetnames[0]] #엑셀파일에서 어느 워크시트를 사용할 것인지 설정한다
  # print(sheet)

  #엑셀 내용을 f(여기선 메모장)에 적는다
  #엑셀의 내용을 셀 단위로 f에 넣고 매 셀 내용마다 줄바꿈을 하여 저장한다
  for row in ws.iter_rows():
    for cell in row:
      if cell.value == None:
        break
      # print(cell.value)
      f.write(cell.value + "\n")
    # print("\n\n\n")
    f.write("\n\n\n")

  f.close()

# wellness_dialog_dataset.xlsx에서 감정    질문 형태로 메모장파일(wellness_dialog_question.txt)을 생성한다
def wellness_question_data():
  root_path = "../data"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_q_output = root_path + "/wellness_dialog_question.txt"

  f = open(wellness_q_output, 'w')

  wb = load_workbook(filename=wellness_file)

  ws = wb[wb.sheetnames[0]]
  # print(sheet)
  # 해당 워크시트에서 행단위로 데이터를 불러와서 f에 적는데 그 행의 첫번째 열(감정)과 두번째 열(유저의 대화)의 셀의 내용만 사용한다 그리고 그 내용 사이에 뛰어쓰기 네번을 넣는다
  for row in ws.iter_rows():
    f.write(row[0].value + "    " + row[1].value + "\n")

  f.close()
# wellness_dialog_dataset.xlsx에서 감정    답변 형태로 wellness_dialog_answer.txt파일을 생성한다.
def wellness_answer_data():
  root_path = "../data"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_a_output = root_path + "/wellness_dialog_answer.txt"

  f = open(wellness_a_output, 'w')
  wb = load_workbook(filename=wellness_file)
  ws = wb[wb.sheetnames[0]]
  # row[2]에는 답변 정보가 있다
  for row in ws.iter_rows():
    if row[2].value == None:
      continue
    else:
      f.write(row[0].value + "    " + row[2].value + "\n")
  f.close()

# chatbot_wellness_data.txt에서 카테고리(감정)정보만 모은 메모장파일(chatbot_wellness_category.txt)을 생성한다
def category_data():
  root_path = "../data"
  data_path = root_path + "/chatbot_wellness_data.txt"
  c_output = root_path + "/chatbot_wellness_category.txt"

  i_f = open(data_path, 'r') 
  o_f = open(c_output, 'w')

  category_count = 0
  flag = True

  cate_dict = []
  i_lines = i_f.readlines() 

  # o_f파일에 감정정보를 a.strip() + "    " + str(category_count) + "\n"형태로 담는다
  for i, data in enumerate(i_lines): #enumerate함수는 (인덱스,내용)의 튜플 형태로 만들어준다(https://www.daleseo.com/python-enumerate/참고) i에 인덱스, data에 내용이 들어가는듯
    tmp = data.split('    ') #뛰어쓰기 네번을 기준으로 내용을 잘라서 리스트를 만든다
    a = tmp[1][:-1] #현재 tmp는 1차원 배열이지만 문자열은 2차원 배열처럼 볼수 있다고 합니다 bread라는 단어를 예로 들면 1행에는 0,1,2,3,4의 인덱스가 들어 있고, 2행에는 b,r,e,a,d 가 들어있습니다.
    q = tmp[0] #tmp[0]는 0,1,2,...의 1행정보이다
    if a not in cate_dict: #cate_dict에 a라는 단어가 들어있지 않으면 cate_dict에 a를 추가하고 o_f파일에 아래 형태로 적는다
      cate_dict.append(a)
      o_f.write(a.strip() + "    " + str(category_count) + "\n")#.strip()은 문자 양옆의 공백을 없앤다 https://codechacha.com/ko/python-string-strip/ 참고
      category_count += 1
  o_f.close()
  i_f.close()

# wellness_dialog_category.txt와 wellness_dialog_question.txt를 기반으로 질문    감정 형태의 메모장 파일(wellness_dialog_for_text_classification.txt)을 생성한다
def wellness_text_classification_data():
  root_path = "../data"
  wellness_category_file = root_path + "/wellness_dialog_category.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"

  cate_file = open(wellness_category_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  text_classfi_file = open(wellness_text_classification_file, 'w')

  category_lines = cate_file.readlines()
  cate_dict = {}
  for line_num, line_data in enumerate(category_lines):
    data = line_data.split('    ')
    cate_dict[data[0]] = data[1][:-1] #data[1][:-1]에 대한 자세한 설명은 4.10폴더를 참고하세요
  print(cate_dict)

  ques_lines = ques_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(ques_lines):
    data = line_data.split('    ')
    # print(data[1]+ "    " + cate_dict[data[0]])
    text_classfi_file.write(data[1][:-1] + "    " + cate_dict[data[0]] + "\n")

  cate_file.close()
  ques_file.close()
  text_classfi_file.close()


# (wellness_dialog_answer.txt)(wellness_dialog_question.txt)를 기반으로 감정이 일치하면 질문    답변 형태의 wellness_dialog_for_autoregressive.txt를 생성
def wellness_dialog_for_autoregressive():
  root_path = "../data"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive.txt"


  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  #모든 질문 데이터에 대하여 감정이 같은 답변을 찾고 일치하면 아래와 같은 형태로 wellness_dialog_for_autoregressive.txt에 저장합니다
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]:
        autoregressive_file.write(ques_data[1][:-1] + "    " + ans_data[1]) #data[1][:-1]에 대한 자세한 설명은 4.10폴더를 참고하세요
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()

#tweeter_dialog_data.txt로 tweeter_dialog_for_autoregressive.txt를 생성
def tweet_data_for_autoregressive():
  root_path = "../data"

  # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
  # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
  file_path = root_path + "/tweeter_dialog_data.txt"
  tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive.txt"

  data_file = open(file_path, 'r')
  tweet_file = open(tweeter_autoregressive_file, 'w')

  data_file_lines = data_file.readlines()
  dialog = ''
  #데이터가 존재하면 dialog에 계속해서 그 데이터를 누적하고 줄바꿈이 나오면 파일에 적는다 요약하면 tweeter_dialog_for_autoregressive.txt파일에 데이터를 저장한다
  for line_num, line_data in enumerate(data_file_lines):
    if line_data == "\n" and dialog != '':#데이터가 줄바꿈일 때(이 줄에 데이터가 존재하지 않을때) dialog에 줄바꿈을 추가하고 dialog를 파일에 적는다
      dialog += "\n"
      tweet_file.write(dialog)
      print(dialog)
      dialog = ''
    elif line_data != "\n": #데이터가 줄바꿈이 아닐때(=데이터가 존재할때) 아래 형태로 dialog에 데이터를 추가한다
      dialog += "<s>" + line_data[:-1] + "</s>"
  data_file.close()
  tweet_file.close()

#wellness_dialog_for_autoregressive.txt로  wellness_dialog_for_autoregressive_train.txt와 wellness_dialog_for_autoregressive_test.txt를 생성
#(질문과 답변이 저장돼있는 파일)을 9:1로 훈련용파일과 테스트용 파일로 나누어 파일을 생성한다
def seperate_wellness_data():
  # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
  # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
  file_path = root_path + "/wellness_dialog_for_autoregressive.txt"
  train_file_path = root_path + "/wellness_dialog_for_autoregressive_train.txt"
  test_file_path = root_path + "/wellness_dialog_for_autoregressive_test.txt"

  sperated_file = open(file_path, 'r')
  train_file = open(train_file_path, 'w')
  test_file = open(test_file_path, 'w')

  sperated_file_lines = sperated_file.readlines()
  ques_dict = {}
  for line_num, line_data in enumerate(sperated_file_lines):
    rand_num = random.randint(0, 10) #0,1,2,...,10의 수 중 랜덤하게 뽑는다
    #랜덤하게 뽑은 수 중 else부분에 해당될 때는 10일 때 밖에 없다. 즉 데이터의 90%는 train파일 10%는 테스트파일로 쓰겠다는 의미
    if rand_num < 10:
      train_file.write(line_data)
    else:
      test_file.write(line_data)

  sperated_file.close()
  train_file.close()
  test_file.close()

#tweeter_dialog_data.txt의 데이터를 변수에 누적하다가 get_kogpt2_tokenizer()로 encode했을 때의 길이가 1024가 넘어가기 전까지의 데이터를 wellness_dialog_for_text_classification.txt에 적고 생성한다. 그리고 토큰화된 누적 데이터의 총 길이를 출력한다
def tweeter_autoregressive_data():
  root_path = "../data"
  tokenizer =get_kogpt2_tokenizer()
  # wellness_autoregressive_file = root_path+"/wellness_dialog_for_autoregressive.txt"
  # wellness_text_classification_file = root_path + "/wellness_dialog_for_text_classification.txt"
  file_path = root_path + "/tweeter_dialog_data.txt"
  tweeter_autoregressive_file = root_path + "/tweeter_dialog_for_autoregressive.txt"

  data_file = open(file_path, 'r')
  tweet_file = open(tweeter_autoregressive_file, 'w')

  data_file_lines = data_file.readlines()
  dialog = ''
  max_len=0

  #데이터가 존재하면 dialog에 계속해서 그 데이터를 누적하고 줄바꿈이 나오면 파일에 적는다 그리고 max_len변수에 누적된 데이터의 길이를 저장한다
  for line_num, line_data in enumerate(data_file_lines):
    if line_data == "\n" and dialog != '':
      dialog += "\n"
      tweet_file.write(dialog)
      print(dialog)
      dialog = ''
    #tmp_data에 계속해서 데이터가 누적되는대 get_kogpt2_tokenizer()로 encode했을 때의 그 길이가 1024를 넘어가면 그 문장부터 파일이 끝날 때까지 누적하지 않고 파일에 적게 된다
    elif line_data != "\n":
      tmp_data = dialog + "<s>" + line_data[:-1] + "</s>"
      if len(tokenizer.encode(tmp_data))>= 1024: 
        continue
      else:
        max_len= max(len(tokenizer.encode(tmp_data)),max_len) #len(tokenizer.encode(tmp_data)와 max_len중 큰 것을 출력한다
        dialog = tmp_data
  print('max_token_length: ', max_len)
  data_file.close()
  tweet_file.close()

# <s>와 같은 토큰을 추가하는 것을 제외하면 wellness_dialog_for_autoregressive()함수와 동일합니다
# wellness_dialog_answer.txt와 wellness_dialog_question.txt를 이용해 아래와 같은 형태로 감정이 일치하는 질문과 답변을 wellness_dialog_for_autoregressive_with_token.txt에 저장합니다
def tweeter_autoregressive_data_with_token():
  root_path = "../data"
  wellness_file = root_path + "/wellness_dialog_dataset.xlsx"
  wellness_answer_file = root_path + "/wellness_dialog_answer.txt"
  wellness_question_file = root_path + "/wellness_dialog_question.txt"
  wellness_autoregressive_file = root_path + "/wellness_dialog_for_autoregressive_with_token.txt"

  answ_file = open(wellness_answer_file, 'r')
  ques_file = open(wellness_question_file, 'r')
  autoregressive_file = open(wellness_autoregressive_file, 'w')

  answ_lines = answ_file.readlines()
  ques_lines = ques_file.readlines()
  ques_dict = {}
  #모든 질문 데이터에 대하여 감정이 같은 답변을 찾고 일치하면 아래와 같은 형태로 wellness_dialog_for_autoregressive_with_token.txt에 저장합니다
  for line_num, line_data in enumerate(ques_lines):
    ques_data = line_data.split('    ')
    for ans_line_num, ans_line_data in enumerate(answ_lines):
      ans_data = ans_line_data.split('    ')
      if ques_data[0] == ans_data[0]: # =감정이 같으면
        autoregressive_file.write("<s>" + ques_data[1][:-1] + "</s><s>" + ans_data[1][:-1] + "</s>\n")
      else:
        continue

  answ_file.close()
  ques_file.close()
  autoregressive_file.close()

#chatbot_data.txt와 wellness_dialog.txt 데이터를 a_chatbot_wellness_data.txt에 적고 생성하는대 그 순서는 chatbot_data.txt가 먼저입니다
#다른 파일들은 모두 이 파일 내의 함수에 의해서 생성되는대 여기서 사용되는 chatbot_data.txt와 wellness_dialog.txt 파일은 출처를 모르겠습니다
def merge_data():
  root_path = "../data"

  chatbot_file = root_path + "/chatbot_data.txt"
  wellness_file = root_path + "/wellness_dialog.txt"

  total_data_file = root_path + "/a_chatbot_wellness_data.txt"

  chatbot_f = open(chatbot_file, 'r')
  wellness_f = open(wellness_file, 'r')
  output_f = open(total_data_file, 'w')

  chatbot_lines = chatbot_f.readlines()
  for line_num, line_data in enumerate(chatbot_lines):
    output_f.write(line_data)

  wellness_lines = wellness_f.readlines()
  for line_num, line_data in enumerate(wellness_lines):
    output_f.write(line_data)

  chatbot_f.close()
  wellness_f.close()
  output_f.close()


if __name__ == "__main__":
  root_path = "../data"
  file_path = root_path + "/chatbot_wellness_data.txt"
  o_path = root_path + "/chatbot_wellness_data_for_autoregressive.txt"

  i_file = open(file_path, 'r')
  o_file = open(o_path, 'w')

  #tmp[1][:-1]에 대한 자세한 설명은 4.10폴더를 참고하세요
  i_lines = i_file.readlines()
  for i, data in enumerate(i_lines):
    tmp = data.split('    ')
    question = tmp[0]
    answer = tmp[1][:-1]
    o_file.write("<s>" + question + "</s><s>" + answer+ "</s>\n")